#!/usr/bin/env python3
import argparse
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook

API_BASE = "https://api.epicentrm.com.ua"
DEFAULT_TARGETS = Path(__file__).with_name("epicenter_target_categories.txt")
DEFAULT_OUTPUT_XLSX = Path(__file__).with_name("epicenter_v2_summary.xlsx")
DEFAULT_OUTPUT_JSON = Path(__file__).with_name("epicenter_v2_snapshot.json")


def log(message: str) -> None:
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {message}", flush=True)


def normalize_title(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip()).casefold()


def read_target_categories(path: Path) -> list[str]:
    lines = path.read_text(encoding="utf-8").splitlines()
    targets: list[str] = []
    for line in lines:
        value = line.strip()
        if not value or value.startswith("#"):
            continue
        targets.append(value)
    return targets


def pick_translation_title(translations: list[dict[str, Any]] | None) -> str:
    translations = translations or []
    for lang in ("ua", "uk", "ru", "en"):
        for tr in translations:
            if (tr.get("languageCode") or "").lower() == lang and tr.get("title"):
                return str(tr["title"]).strip()
    for tr in translations:
        if tr.get("title"):
            return str(tr["title"]).strip()
    return ""


class EpicenterClient:
    def __init__(self, token: str, timeout: int = 60) -> None:
        self.session = requests.Session()
        self.session.headers.update(
            {
                "Authorization": f"Bearer {token}",
                "Accept": "application/json",
            }
        )
        self.timeout = timeout

    def _get(self, path: str, params: Any = None) -> dict[str, Any]:
        url = f"{API_BASE}{path}"
        response = self.session.get(url, params=params or {}, timeout=self.timeout)
        response.raise_for_status()
        data = response.json()
        if not isinstance(data, dict):
            raise RuntimeError(f"Unexpected response for {path}: {type(data)}")
        return data

    def get_paged(
        self,
        path: str,
        params: dict[str, Any] | None = None,
        label: str | None = None,
    ) -> list[dict[str, Any]]:
        params = dict(params or {})
        params.setdefault("page", 1)
        params.setdefault("per-page", 200)
        items: list[dict[str, Any]] = []

        while True:
            data = self._get(path, params=params)
            page_items = data.get("items") or []
            if not isinstance(page_items, list):
                raise RuntimeError(f"Invalid items format for {path}")
            items.extend(page_items)

            page = int(data.get("page") or params["page"])
            pages = int(data.get("pages") or page)
            if label:
                log(f"{label}: page {page}/{pages}, collected {len(items)}")
            if page >= pages:
                break
            params["page"] = page + 1

        return items


def chunked(values: list[str], size: int) -> list[list[str]]:
    return [values[i : i + size] for i in range(0, len(values), size)]


def write_workbook(
    output_xlsx: Path,
    target_categories: list[str],
    matched_categories: list[dict[str, Any]],
    attributes_rows: list[dict[str, Any]],
    options_rows: list[dict[str, Any]],
) -> None:
    wb = Workbook()

    ws_categories = wb.active
    ws_categories.title = "categories"
    ws_categories.append(
        [
            "target_title",
            "matched",
            "category_code",
            "category_title",
            "attribute_set_codes",
        ]
    )

    matched_by_target: dict[str, list[dict[str, Any]]] = {}
    for c in matched_categories:
        matched_by_target.setdefault(c["target_title"], []).append(c)

    for title in target_categories:
        rows = matched_by_target.get(title) or []
        if not rows:
            ws_categories.append([title, "no", "", "", ""])
            continue
        for item in rows:
            ws_categories.append(
                [
                    title,
                    "yes",
                    item.get("category_code", ""),
                    item.get("category_title", ""),
                    ",".join(item.get("attribute_set_codes") or []),
                ]
            )

    ws_attr = wb.create_sheet("attributes")
    ws_attr.append(
        [
            "category_title",
            "category_code",
            "attribute_set_code",
            "attribute_title",
            "attribute_code",
            "attribute_type",
            "is_filter",
            "is_required",
            "is_system",
            "is_model",
            "prefix",
            "suffix",
        ]
    )
    for row in attributes_rows:
        ws_attr.append(
            [
                row.get("category_title", ""),
                row.get("category_code", ""),
                row.get("attribute_set_code", ""),
                row.get("attribute_title", ""),
                row.get("attribute_code", ""),
                row.get("attribute_type", ""),
                row.get("is_filter", False),
                row.get("is_required", False),
                row.get("is_system", False),
                row.get("is_model", False),
                row.get("prefix", ""),
                row.get("suffix", ""),
            ]
        )

    ws_options = wb.create_sheet("options")
    ws_options.append(
        [
            "category_title",
            "category_code",
            "attribute_set_code",
            "attribute_title",
            "attribute_code",
            "option_title",
            "option_code",
        ]
    )
    for row in options_rows:
        ws_options.append(
            [
                row.get("category_title", ""),
                row.get("category_code", ""),
                row.get("attribute_set_code", ""),
                row.get("attribute_title", ""),
                row.get("attribute_code", ""),
                row.get("option_title", ""),
                row.get("option_code", ""),
            ]
        )

    wb.save(output_xlsx)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Export Epicenter V2 category/attribute/options reference into editable summary table"
    )
    parser.add_argument("--targets", type=Path, default=DEFAULT_TARGETS, help="Path to target categories txt")
    parser.add_argument("--output-xlsx", type=Path, default=DEFAULT_OUTPUT_XLSX, help="Result xlsx path")
    parser.add_argument("--output-json", type=Path, default=DEFAULT_OUTPUT_JSON, help="Raw snapshot json path")
    parser.add_argument("--token", default=os.environ.get("EPICENTER_API_TOKEN", ""), help="Epicenter API token")
    args = parser.parse_args()

    if not args.targets.exists():
        raise SystemExit(f"Targets file not found: {args.targets}")
    if not args.token:
        raise SystemExit("EPICENTER_API_TOKEN is not set. Pass --token or export env var.")

    targets = read_target_categories(args.targets)
    target_norm = {normalize_title(x): x for x in targets}

    client = EpicenterClient(args.token)
    log("Step 1/5: Fetch categories from /v2/pim/categories")
    categories = client.get_paged("/v2/pim/categories", label="Categories")
    log(f"Categories fetched: {len(categories)}")

    log("Step 2/5: Match target category titles")
    category_matches: list[dict[str, Any]] = []
    unmatched = set(targets)
    for category in categories:
        title = pick_translation_title(category.get("translations"))
        norm_title = normalize_title(title)
        target_title = target_norm.get(norm_title)
        if not target_title:
            continue
        as_codes = [x.get("code", "") for x in (category.get("attributeSets") or []) if x.get("code")]
        category_matches.append(
            {
                "target_title": target_title,
                "category_code": str(category.get("code", "")).strip(),
                "category_title": title,
                "attribute_set_codes": as_codes,
            }
        )
        if target_title in unmatched:
            unmatched.remove(target_title)

    attribute_set_codes = sorted(
        {
            code
            for item in category_matches
            for code in (item.get("attribute_set_codes") or [])
            if code
        }
    )
    log(f"Matched categories: {len(category_matches)}; unique attribute sets: {len(attribute_set_codes)}")

    log("Step 3/5: Fetch attribute sets")
    attr_sets_map: dict[str, dict[str, Any]] = {}
    chunks = chunked(attribute_set_codes, 50)
    for idx, chunk in enumerate(chunks, start=1):
        log(f"Attribute sets batch {idx}/{len(chunks)}: requesting {len(chunk)} codes")
        params: list[tuple[str, str]] = [("page", "1"), ("per-page", "200")]
        for code in chunk:
            params.append(("filter[codes][]", code))
        data = client._get("/v2/pim/attribute-sets", params=params)
        page_items = data.get("items") or []
        log(f"Attribute sets batch {idx}/{len(chunks)}: received {len(page_items)} items")
        for item in data.get("items") or []:
            code = str(item.get("code", "")).strip()
            if code:
                attr_sets_map[code] = item
    log(f"Attribute sets loaded: {len(attr_sets_map)}")

    options_cache: dict[tuple[str, str], list[dict[str, Any]]] = {}
    attributes_rows: list[dict[str, Any]] = []
    options_rows: list[dict[str, Any]] = []
    options_requests = 0

    log("Step 4/5: Build rows and fetch options for select/multiselect attributes")
    for category in category_matches:
        cat_title = category["category_title"]
        cat_code = category["category_code"]
        log(f"Category {cat_title} ({cat_code}) with {len(category['attribute_set_codes'])} attribute sets")
        for as_code in category["attribute_set_codes"]:
            attr_set = attr_sets_map.get(as_code) or {}
            for attr in (attr_set.get("attributes") or []):
                attr_code = str(attr.get("code", "")).strip()
                attr_type = str(attr.get("type", "")).strip()
                translations = attr.get("translations") or []
                attr_title = pick_translation_title(translations)

                prefix = ""
                suffix = ""
                for tr in translations:
                    if (tr.get("languageCode") or "").lower() in {"ua", "uk", "ru"}:
                        prefix = tr.get("prefix") or prefix
                        suffix = tr.get("suffix") or suffix

                attributes_rows.append(
                    {
                        "category_title": cat_title,
                        "category_code": cat_code,
                        "attribute_set_code": as_code,
                        "attribute_title": attr_title,
                        "attribute_code": attr_code,
                        "attribute_type": attr_type,
                        "is_filter": bool(attr.get("isFilter")),
                        "is_required": bool(attr.get("isRequired")),
                        "is_system": bool(attr.get("isSystem")),
                        "is_model": bool(attr.get("isModel")),
                        "prefix": prefix,
                        "suffix": suffix,
                    }
                )

                if attr_type not in {"select", "multiselect"}:
                    continue
                cache_key = (as_code, attr_code)
                if cache_key not in options_cache:
                    path = f"/v2/pim/attribute-sets/{as_code}/attributes/{attr_code}/options"
                    options_requests += 1
                    options_cache[cache_key] = client.get_paged(
                        path,
                        label=f"Options {options_requests}: {as_code}/{attr_code}",
                    )

                for opt in options_cache[cache_key]:
                    options_rows.append(
                        {
                            "category_title": cat_title,
                            "category_code": cat_code,
                            "attribute_set_code": as_code,
                            "attribute_title": attr_title,
                            "attribute_code": attr_code,
                            "option_title": pick_translation_title(opt.get("translations")),
                            "option_code": str(opt.get("code", "")).strip(),
                        }
                    )

    log("Step 5/5: Save workbook and snapshot")
    write_workbook(args.output_xlsx, targets, category_matches, attributes_rows, options_rows)
    snapshot = {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "targets": targets,
        "unmatched_targets": sorted(unmatched),
        "matched_categories": category_matches,
        "attribute_set_codes": attribute_set_codes,
        "attributes_rows_count": len(attributes_rows),
        "options_rows_count": len(options_rows),
    }
    args.output_json.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")

    log(f"Targets: {len(targets)}")
    log(f"Matched categories: {len(category_matches)}")
    log(f"Unmatched categories: {len(unmatched)}")
    log(f"Attribute sets: {len(attribute_set_codes)}")
    log(f"Attributes rows: {len(attributes_rows)}")
    log(f"Options rows: {len(options_rows)}")
    log(f"Saved: {args.output_xlsx}")
    log(f"Saved: {args.output_json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
